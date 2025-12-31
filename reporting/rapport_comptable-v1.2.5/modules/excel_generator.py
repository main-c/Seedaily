#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module de génération du fichier Excel

Ce module gère:
- La création du classeur Excel multi-feuilles
- L'insertion des données dans chaque feuille
- L'application du formatage (couleurs, bordures, alignements)
- L'insertion des formules Excel pour les calculs dynamiques
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
import logging
from typing import Dict

from utils.exceptions import ExcelGenerationError

logger = logging.getLogger(__name__)


def create_workbook(
    df_grand_livre: pd.DataFrame,
    df_balance: pd.DataFrame,
    bilan: Dict,
    compte_resultat: Dict,
    sig: Dict,
    client_code: str = None,
) -> Workbook:
    """
    Crée un nouveau classeur Excel avec toutes les feuilles

    Args:
        df_grand_livre: DataFrame du Grand Livre
        df_balance: DataFrame de la Balance
        bilan: Dictionnaire du bilan
        compte_resultat: Dictionnaire du compte de résultat
        sig: Dictionnaire des SIG
        client_code: Code du client pour utiliser son mapping spécifique (optionnel)

    Returns:
        Workbook openpyxl
    """
    logger.info("Création du classeur Excel")
    if client_code:
        logger.info(f"Utilisation du mapping pour le client: {client_code}")

    try:
        wb = Workbook()

        # Supprimer la feuille par défaut
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Ajouter les différentes feuilles
        add_grand_livre_sheet(wb, df_grand_livre)
        add_balance_sheet(wb, df_balance)

        # Détecter le format du bilan (ancien format avec clés fixes vs nouveau format avec listes)
        # Nouveau format: bilan['actif'] est une liste de dicts avec 'poste' et 'montant'
        # Ancien format: bilan['actif'] est un dict avec des clés comme 'immo_incorp', 'stocks', etc.
        if isinstance(bilan.get("actif"), list):
            logger.info("Détection du nouveau format de bilan (depuis correspondances)")
            add_bilan_sheet_from_mapping(wb, bilan)
        else:
            logger.info("Détection de l'ancien format de bilan")
            add_bilan_sheet(wb, bilan)

        # Détecter le format du CR (ancien format avec clés fixes vs nouveau format avec listes)
        if isinstance(compte_resultat.get("charges"), list):
            logger.info("Détection du nouveau format de CR (depuis correspondances)")
            add_compte_resultat_sheet_from_mapping(wb, compte_resultat)
        else:
            logger.info("Détection de l'ancien format de CR")
            add_compte_resultat_sheet(wb, compte_resultat)

        add_sig_sheet(wb, sig)
        add_suivi_activite_sheet(wb, df_grand_livre, client_code=client_code)

        logger.info(f"Classeur créé avec {len(wb.sheetnames)} feuilles")

        return wb

    except Exception as e:
        logger.error(f"Erreur lors de la création du classeur: {e}")
        raise ExcelGenerationError(f"Impossible de créer le classeur Excel: {e}")


def add_grand_livre_sheet(wb: Workbook, df: pd.DataFrame):
    """Ajoute la feuille Grand Livre"""
    logger.info("Ajout de la feuille Grand Livre")

    ws = wb.create_sheet("GL BI SEP")

    # Ajouter les en-têtes
    headers = [
        "N° Compte",
        "Date",
        "Journal",
        "N° Pièce",
        "Libellé",
        "Lettrage",
        "Débit",
        "Crédit",
        "Solde",
    ]
    ws.append(headers)

    # Appliquer le style d'en-tête
    apply_header_style(ws, 1, len(headers))

    # Ajouter les données
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=False), start=2
    ):
        ws.append(row)

    # Appliquer le formatage des nombres
    apply_number_format(ws, "G", 2, len(df) + 1)  # Débit
    apply_number_format(ws, "H", 2, len(df) + 1)  # Crédit
    apply_number_format(ws, "I", 2, len(df) + 1)  # Solde

    # Ajouter les 3 lignes de totaux (comme dans le fichier manuel)
    last_data_row = len(df) + 1

    # Ligne 1: Totaux des colonnes Débit et Crédit
    total_row_1 = last_data_row + 1
    ws.append(
        [
            "",
            "",
            "",
            "",
            "",
            "",
            f"=SUM(G2:G{last_data_row})",
            f"=SUM(H2:H{last_data_row})",
            "",
        ]
    )
    apply_number_format(ws, "G", total_row_1, total_row_1)
    apply_number_format(ws, "H", total_row_1, total_row_1)

    # Ligne 2: Ligne vide (séparateur)
    ws.append(["", "", "", "", "", "", "", "", ""])

    # Ligne 3: Total final (balance = Crédit - Débit)
    # Dans le fichier manuel, cette valeur est dans la colonne Crédit (H)
    total_row_3 = last_data_row + 3
    ws.append(["", "", "", "", "", "", "", f"=H{total_row_1}-G{total_row_1}", ""])
    apply_number_format(ws, "H", total_row_3, total_row_3)

    # Appliquer un style de mise en évidence pour les lignes de totaux
    apply_total_style(ws, total_row_1, len(headers))
    apply_total_style(ws, total_row_3, len(headers))

    # Ajuster la largeur des colonnes
    adjust_column_widths(ws)


def add_balance_sheet(wb: Workbook, df: pd.DataFrame):
    """
    Ajoute la feuille Balance Générale (BG BI SEP)

    Cette feuille est la source fondamentale pour toutes les autres analyses:
    - Bilan: utilise les soldes débiteurs (Actif) et créditeurs (Passif)
    - Compte de Résultat: extrait les comptes classe 6 (Charges) et 7 (Produits)
    - SIG: calcule les soldes intermédiaires à partir des sous-classes
    """
    logger.info("Ajout de la feuille Balance")

    ws = wb.create_sheet("BG BI SEP")

    # Ajouter 2 lignes vides au début (comme dans le fichier manuel)
    ws.append(["", "", "", "", "", ""])
    ws.append(["", "", "", "", "", ""])

    # Ajouter les en-têtes explicites à la ligne 3
    headers = [
        "N° Compte",
        "Libellé",
        "Total Débit",
        "Total Crédit",
        "Solde Débiteur",
        "Solde Créditeur",
    ]
    ws.append(headers)

    # Appliquer le style d'en-tête
    apply_header_style(ws, 3, len(headers))

    # Ajouter les données avec formules pour les soldes
    # Les colonnes du DataFrame sont: compte, libelle, total_debit, total_credit, solde_debiteur, solde_crediteur
    for r_idx, (_, row) in enumerate(df.iterrows(), start=4):
        compte = row["compte"]
        libelle = row["libelle"]
        debit = row["total_debit"]
        credit = row["total_credit"]

        # Calcul des soldes avec formules Excel
        # Solde Débiteur = MAX(0, Débit - Crédit)
        # Solde Créditeur = MAX(0, Crédit - Débit)
        solde_debiteur = f"=MAX(0,C{r_idx}-D{r_idx})"
        solde_crediteur = f"=MAX(0,D{r_idx}-C{r_idx})"

        ws.append([compte, libelle, debit, credit, solde_debiteur, solde_crediteur])

    # Appliquer le formatage des nombres
    last_data_row = 3 + len(df)
    apply_number_format(ws, "C", 4, last_data_row)  # Total Débit
    apply_number_format(ws, "D", 4, last_data_row)  # Total Crédit
    apply_number_format(ws, "E", 4, last_data_row)  # Solde Débiteur
    apply_number_format(ws, "F", 4, last_data_row)  # Solde Créditeur

    # Ajouter une ligne de total
    total_row = last_data_row + 1
    ws.append(
        [
            "",
            "TOTAL GÉNÉRAL",
            f"=SUM(C4:C{last_data_row})",
            f"=SUM(D4:D{last_data_row})",
            f"=SUM(E4:E{last_data_row})",
            f"=SUM(F4:F{last_data_row})",
        ]
    )

    # Appliquer le formatage pour la ligne de total
    apply_number_format(ws, "C", total_row, total_row)
    apply_number_format(ws, "D", total_row, total_row)
    apply_number_format(ws, "E", total_row, total_row)
    apply_number_format(ws, "F", total_row, total_row)

    # Style pour la ligne de total
    apply_total_style(ws, total_row, len(headers))

    # Ajuster la largeur des colonnes
    adjust_column_widths(ws)


def add_bilan_sheet_from_mapping(wb: Workbook, bilan: Dict):
    """
    Ajoute la feuille Bilan Synthèse en respectant exactement le format du modèle client.

    Format attendu (selon SUIVI D'ACTIVITE BIMMO SEP.xlsx):
    - Ligne 1-2: En-têtes avec fond bleu (#366092) et texte blanc centré
    - Lignes 3-29: Données avec bordures, police EYInterstate/Arial taille 11
    - Ligne 30: Totaux avec fond gris (#D9E1F2) et police en gras
    - Largeurs de colonnes spécifiques: A=50.5, B=4.5, C=18, D=34.8, E=17.5, F=18
    - Format nombres: #,##0.00 aligné à droite
    """
    logger.info("Ajout de la feuille Bilan Synthèse (format client)")

    ws = wb.create_sheet("BILAN SYNTH")
    ws.sheet_view.showGridLines = False

    # Palette de couleurs professionnelle
    header_color = "4472C4"  # Bleu corporate
    subheader_color = "D9E1F2"  # Bleu très pâle
    total_color = "8EA9DB"  # Bleu intermédiaire
    section_color = "E6EBF5"  # Bleu très très pâle

    # Définition des styles
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    subheader_font = Font(name="Calibri", size=11, bold=True)
    data_font = Font(name="Calibri", size=11)
    total_font = Font(name="Calibri", size=11, bold=True)

    # En-têtes et alignements
    header_alignment = Alignment(horizontal="center", vertical="center")
    label_alignment = Alignment(horizontal="left", vertical="center")
    amount_alignment = Alignment(horizontal="right", vertical="center")

    # Bordures
    no_border = Border()
    bottom_border = Border(bottom=Side(style="thin", color="BFBFBF"))
    top_border = Border(top=Side(style="thin", color="BFBFBF"))
    top_bottom_border = Border(
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    outer_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    # Remplissages
    header_fill = PatternFill(
        start_color=header_color, end_color=header_color, fill_type="solid"
    )
    subheader_fill = PatternFill(
        start_color=subheader_color, end_color=subheader_color, fill_type="solid"
    )
    total_fill = PatternFill(
        start_color=total_color, end_color=total_color, fill_type="solid"
    )
    section_fill = PatternFill(
        start_color=section_color, end_color=section_color, fill_type="solid"
    )

    # Définir les largeurs de colonnes
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 2  # Espace entre Actif et Passif
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 18

    # Préparer les données
    actif_data = bilan["actif"]
    passif_data = bilan["passif"]

    # ---- EN-TÊTE DU BILAN ----
    row = 1
    ws.append(["ACTIF", "", "", "PASSIF", ""])

    # Appliquer le style à l'en-tête
    for col in ["A", "B", "D", "E"]:
        cell = ws[f"{col}{row}"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = no_border

    # Colonne C (séparation)
    cell = ws[f"C{row}"]
    cell.border = no_border

    row += 1
    ws.append(["", "", "", "", ""])  # Ligne d'espacement
    row += 1

    # ---- SECTIONS DU BILAN ----
    # Liste des sections pour l'organisation
    sections = [
        {
            "title": "IMMOBILISATIONS",
            "actif": [
                poste for poste in actif_data if "IMMOBILISATION" in poste["poste"]
            ],
            "passif": [
                poste
                for poste in passif_data
                if poste["poste"] in ["CAPITAL SOCIAL", "RAN", "RESULTAT NET"]
            ],
        },
        {
            "title": "STOCKS ET CRÉANCES",
            "actif": [
                poste
                for poste in actif_data
                if any(
                    term in poste["poste"]
                    for term in ["STOCK", "CREANCE", "FOURNISSEUR", "DEBITEUR"]
                )
            ],
            "passif": [
                poste
                for poste in passif_data
                if any(
                    term in poste["poste"]
                    for term in ["DETTE", "FOURNISSEUR", "CREDITEUR"]
                )
            ],
        },
        {
            "title": "TRÉSORERIE",
            "actif": [
                poste
                for poste in actif_data
                if any(
                    term in poste["poste"]
                    for term in ["BANQUE", "CAISSE", "CHEQUE", "MONNAIE", "VIREMENT"]
                )
            ],
            "passif": [poste for poste in passif_data if "DECOUVERT" in poste["poste"]],
        },
    ]

    # Créer chaque section
    for section in sections:
        # Titre de la section
        ws.append([section["title"], "", "", section["title"], ""])
        for col in ["A", "B", "D", "E"]:
            cell = ws[f"{col}{row}"]
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.border = bottom_border
            cell.alignment = label_alignment
        row += 1

        # Contenu de la section
        max_rows = max(len(section["actif"]), len(section["passif"]))
        for i in range(max_rows):
            actif_poste = (
                section["actif"][i]
                if i < len(section["actif"])
                else {"poste": "", "montant": ""}
            )
            passif_poste = (
                section["passif"][i]
                if i < len(section["passif"])
                else {"poste": "", "montant": ""}
            )

            ws.append(
                [
                    actif_poste["poste"],
                    actif_poste["montant"],
                    "",
                    passif_poste["poste"],
                    passif_poste["montant"],
                ]
            )

            # Style pour l'Actif
            cell_label = ws[f"A{row}"]
            cell_label.font = data_font
            cell_label.alignment = label_alignment

            cell_amount = ws[f"B{row}"]
            cell_amount.font = data_font
            cell_amount.alignment = amount_alignment

            if actif_poste["montant"] != "":
                if isinstance(actif_poste["montant"], (int, float)) and actif_poste[
                    "montant"
                ] == int(actif_poste["montant"]):
                    cell_amount.number_format = "# ##0"
                else:
                    cell_amount.number_format = "# ##0,00"

            # Style pour le Passif
            cell_label = ws[f"D{row}"]
            cell_label.font = data_font
            cell_label.alignment = label_alignment

            cell_amount = ws[f"E{row}"]
            cell_amount.font = data_font
            cell_amount.alignment = amount_alignment

            if passif_poste["montant"] != "":
                if isinstance(passif_poste["montant"], (int, float)) and passif_poste[
                    "montant"
                ] == int(passif_poste["montant"]):
                    cell_amount.number_format = "# ##0"
                else:
                    cell_amount.number_format = "# ##0,00"

            # Ajouter un léger fond de couleur alternée pour faciliter la lecture
            if i % 2 == 1:
                for col in ["A", "B", "D", "E"]:
                    ws[f"{col}{row}"].fill = PatternFill(
                        start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                    )

            row += 1

        # Ajouter une ligne d'espacement après chaque section
        ws.append(["", "", "", "", ""])
        row += 1

    # ---- TOTAUX ----
    ws.append(
        ["TOTAL ACTIF", bilan["total_actif"], "", "TOTAL PASSIF", bilan["total_passif"]]
    )

    # Appliquer le style aux totaux
    for col in ["A", "B", "D", "E"]:
        cell = ws[f"{col}{row}"]
        cell.fill = total_fill
        cell.font = total_font
        cell.border = top_bottom_border

        if col in ["A", "D"]:
            cell.alignment = label_alignment
        else:  # col in ['B', 'E']
            cell.alignment = amount_alignment

            # Formatage des totaux
            if col == "B":
                if isinstance(bilan["total_actif"], (int, float)) and bilan[
                    "total_actif"
                ] == int(bilan["total_actif"]):
                    cell.number_format = "# ##0"
                else:
                    cell.number_format = "# ##0,00"
            else:  # col == 'E'
                if isinstance(bilan["total_passif"], (int, float)) and bilan[
                    "total_passif"
                ] == int(bilan["total_passif"]):
                    cell.number_format = "# ##0"
                else:
                    cell.number_format = "# ##0,00"

    # Colonne C (séparation)
    cell = ws[f"C{row}"]
    cell.border = no_border

    # Ajouter une note de pied de page si nécessaire
    row += 2
    ws.append(
        ["Bilan synthétique préparé pour présentation PowerPoint", "", "", "", ""]
    )
    cell = ws[f"A{row}"]
    cell.font = Font(name="Calibri", size=9, italic=True)

    # Vérification de l'équilibre du bilan
    if bilan["total_actif"] == bilan["total_passif"]:
        ws.append(["Bilan équilibré", "", "", "", ""])
    else:
        ws.append(["Attention: Bilan non équilibré", "", "", "", ""])
    cell = ws[f"A{row+1}"]
    cell.font = Font(name="Calibri", size=9, italic=True, color="666666")


def add_bilan_sheet(wb: Workbook, bilan: Dict):
    """
    Ajoute la feuille Bilan Synthétique par catégories comptables

    Format: Actif à gauche (colonnes A-C), Passif à droite (colonnes D-F)
    Conforme au format SYSCOHADA
    """
    logger.info("Ajout de la feuille Bilan Synthèse")

    ws = wb.create_sheet("BILAN SYNTH")

    # Masquer le quadrillage des cellules
    ws.sheet_view.showGridLines = False

    # Appliquer une police professionnelle pour toute la feuille
    default_font = Font(name="Calibri", size=11)

    actif = bilan["actif"]
    passif = bilan["passif"]

    # Ligne 1: Headers ACTIF et PASSIF
    ws.append(["Actif", "", "", "Passif", "", ""])
    apply_header_style(ws, 1, 6)

    ws.append([])  # Ligne vide

    # Ligne 3: ACTIF IMMOBILISÉ | CAPITAUX PROPRES
    ws.append(
        [
            " IMMOBILISATIONS INCORP., VALEURS BRUTES ",
            "",
            actif["immo_incorp"],
            " CAPITAUX PROPRES ",
            "",
            passif["capital"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 5: (vide) | RAN
    ws.append(["", "", "", "RAN", "", passif["ran"]])

    # Ligne 6: IMMOBILISATIONS CORP | (vide)
    ws.append(
        [" IMMOBILISATIONS CORP VALEURS BRUTES", "", actif["immo_corp"], "", "", ""]
    )

    # Ligne 7: (vide) | RESULTAT NET
    ws.append(["", "", "", "RESULTAT NET", "", passif["resultat_net"]])

    # Ligne 8: IMMOBILISATIONS FINANCIERES | (vide)
    ws.append(
        [" IMMOBILISATIONS FINANCIERES", "", actif["immo_financieres"], "", "", ""]
    )

    ws.append([])  # Ligne vide

    # Ligne 10: AMORTISSEMENTS | (vide)
    ws.append([" AMORTISSEMENTS DES IMMO.", "", actif["amortissements"], "", "", ""])

    # Ligne 11: (vide) | DETTES FINANCIÈRES
    ws.append(["", "", "", "DETTES FINANCIERES ", "", passif["dettes_financieres"]])

    # Ligne 12: STOCK | (vide)
    ws.append([" STOCK", "", actif["stocks"], "", "", ""])

    ws.append([])  # Ligne vide
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])

    # Ligne 18: CREANCES CLIENTS | FOURNISSEURS
    ws.append(
        [
            "CREANCES CLIENTS ",
            "",
            actif["creances_clients"],
            " FOURNIS. & COMPTES RATTACHEES ",
            "",
            passif["dettes_fournisseurs"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 20: (vide) | DETTES FISCALES & SOCIALES
    ws.append(
        [
            "",
            "",
            "",
            "DETTES FISCALES & SOCIALES",
            "",
            passif["dettes_fiscales_sociales"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 22: AUTRES CREANCES | AUTRES DETTES
    ws.append(
        [
            " AUTRES CREANCES",
            "",
            actif["autres_creances"],
            " AUTRES DETTES ",
            "",
            passif["autres_dettes"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 24: DEBITEURS DIVERS | CREDITEURS DIVERS
    ws.append(
        [
            " DEBITEURS DIVERS",
            "",
            actif["debiteurs_divers"],
            "CREDITEURS DIVERS",
            "",
            passif["crediteurs_divers"],
        ]
    )

    ws.append([])  # Ligne vide
    ws.append([])
    ws.append([])

    # Ligne 28: BANQUES | (vide)
    ws.append([" BANQUES ", "", actif["banques"], "", "", ""])

    # Ligne 29: CAISSE | DECOUVERTS
    ws.append([" CAISSE", "", actif["caisse"], " DECOUVERTS", "", passif["decouvert"]])

    # Ligne 30: TOTAUX
    ws.append(
        [
            "Total Actif",
            "",
            bilan["total_actif"],
            "Total Passif",
            "",
            bilan["total_passif"],
        ]
    )
    apply_total_style(ws, 30, 6)

    # Appliquer le formatage des nombres
    for row in [3, 5, 6, 7, 8, 10, 11, 12, 18, 20, 22, 24, 28, 29, 30]:
        apply_number_format(ws, "C", row, row)  # Colonne Actif
        apply_number_format(ws, "F", row, row)  # Colonne Passif

    # Appliquer la police Calibri 11 à toutes les cellules
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            if (
                cell.font.name != "Calibri" or not cell.font.bold
            ):  # Ne pas modifier les headers
                cell.font = default_font

    # Ajuster la largeur des colonnes
    adjust_column_widths(ws)

    # Garantir une largeur minimale pour les colonnes de montants
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width, 18)
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width, 18)


def add_compte_resultat_sheet_from_mapping(wb: Workbook, compte_resultat: Dict):
    """
    Compte de Résultat synthétique selon SYSCOHADA avec totaux équilibrés
    Structure: Sous-totaux par nature (Exploitation, Financier, HAO) + Résultat avant totaux
    """
    logger.info("Ajout de la feuille CR synthétique SYSCOHADA avec totaux équilibrés")

    ws = wb.create_sheet("CR SYNTH")
    ws.sheet_view.showGridLines = False

    # En-têtes
    header_row = ["CHARGES", "MONTANTS", "PRODUITS", "MONTANTS"]
    ws.append(header_row)

    # Style pour les en-têtes
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Récupérer les données
    charges_data = compte_resultat.get("charges", [])
    produits_data = compte_resultat.get("produits", [])
    total_charges = compte_resultat.get("total_charges", 0)
    total_produits = compte_resultat.get("total_produits", 0)
    resultat = compte_resultat.get("resultat", 0)

    # Séparer les données par nature (Exploitation, Financier, HAO)
    charges_exploitation = []
    charges_financieres = []
    charges_hao = []
    impot_societes = 0

    produits_exploitation = []
    produits_financiers = []
    produits_hao = []

    # Classifier les charges
    for charge in charges_data:
        poste = charge.get("poste", "")
        montant = charge.get("montant", 0)

        if any(
            kw in poste.lower()
            for kw in [
                "frais financier",
                "interet",
                "escompte accord",
                "perte de change",
            ]
        ):
            charges_financieres.append(charge)
        elif (
            "hors activité" in poste.lower()
            or "hao" in poste.lower()
            or "valeur comptable" in poste.lower()
        ):
            charges_hao.append(charge)
        elif "impot" in poste.lower() and "societe" in poste.lower():
            impot_societes = montant
        else:
            charges_exploitation.append(charge)

    # Classifier les produits
    for produit in produits_data:
        poste = produit.get("poste", "")

        if any(
            kw in poste.lower()
            for kw in ["revenu financier", "escompte obtenu", "gain de change"]
        ):
            produits_financiers.append(produit)
        elif (
            "hors activité" in poste.lower()
            or "hao" in poste.lower()
            or "cession" in poste.lower()
        ):
            produits_hao.append(produit)
        else:
            produits_exploitation.append(produit)

    # Écriture des données EXPLOITATION
    max_rows_exploit = max(len(charges_exploitation), len(produits_exploitation))

    for i in range(max_rows_exploit):
        row_data = ["", "", "", ""]

        if i < len(charges_exploitation):
            charge = charges_exploitation[i]
            row_data[0] = charge.get("poste", "")
            row_data[1] = charge.get("montant", 0)

        if i < len(produits_exploitation):
            produit = produits_exploitation[i]
            row_data[2] = produit.get("poste", "")
            row_data[3] = produit.get("montant", 0)

        ws.append(row_data)

        # Formatage des montants
        for col in [2, 4]:
            cell = ws.cell(row=ws.max_row, column=col)
            if cell.value and cell.value != "":
                cell.number_format = "# ##0"
                cell.alignment = Alignment(horizontal="right")

    # Ligne vide
    ws.append([""] * 4)

    # Données FINANCIÈRES (sans sous-total, juste les lignes)
    max_rows_fin = (
        max(len(charges_financieres), len(produits_financiers))
        if (charges_financieres or produits_financiers)
        else 0
    )
    for i in range(max_rows_fin):
        row_data = ["", "", "", ""]

        if i < len(charges_financieres):
            charge = charges_financieres[i]
            row_data[0] = charge.get("poste", "")
            row_data[1] = charge.get("montant", 0)

        if i < len(produits_financiers):
            produit = produits_financiers[i]
            row_data[2] = produit.get("poste", "")
            row_data[3] = produit.get("montant", 0)

        ws.append(row_data)

        for col in [2, 4]:
            cell = ws.cell(row=ws.max_row, column=col)
            if cell.value and cell.value != "":
                cell.number_format = "# ##0"
                cell.alignment = Alignment(horizontal="right")

    # Données HAO (sans sous-total, juste les lignes)
    max_rows_hao = (
        max(len(charges_hao), len(produits_hao)) if (charges_hao or produits_hao) else 0
    )
    for i in range(max_rows_hao):
        row_data = ["", "", "", ""]

        if i < len(charges_hao):
            charge = charges_hao[i]
            row_data[0] = charge.get("poste", "")
            row_data[1] = charge.get("montant", 0)

        if i < len(produits_hao):
            produit = produits_hao[i]
            row_data[2] = produit.get("poste", "")
            row_data[3] = produit.get("montant", 0)

        ws.append(row_data)

        for col in [2, 4]:
            cell = ws.cell(row=ws.max_row, column=col)
            if cell.value and cell.value != "":
                cell.number_format = "# ##0"
                cell.alignment = Alignment(horizontal="retight")

    # Impôt sur les sociétés (si présent)
    if impot_societes > 0:
        ws.append(["Impot sur les sociétés", impot_societes, "", ""])
        cell = ws.cell(row=ws.max_row, column=2)
        cell.number_format = "# ##0"
        cell.alignment = Alignment(horizontal="right")

    # Ligne RÉSULTAT (avant les totaux) - format 4 colonnes
    ws.append([""] * 4)

    # Sous-total EXPLOITATION
    total_charges_exploit = sum(c.get("montant", 0) for c in charges_exploitation)
    total_produits_exploit = sum(p.get("montant", 0) for p in produits_exploitation)

    ws.append(
        [
            "Charges d'exploitation",
            total_charges_exploit,
            "Produits d'exploitation",
            total_produits_exploit,
        ]
    )
    subtotal_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    for col in range(1, 5):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = subtotal_fill
        if col in [2, 4]:
            cell.number_format = "# ##0"
            cell.alignment = Alignment(horizontal="right")

    result_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )

    if resultat >= 0:
        # Bénéfice (Net Profit +) - Colonne A: "Resultat - Net Profit (+)", Colonne B: montant
        ws.append(["Resultat - Net Profit (+)", resultat, "Resultat - Net Loss (-)", 0])
        # Formater
        for col in [1, 2, 3, 4]:
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = result_fill
            cell.font = Font(bold=True)
            if col in [2, 4]:  # Montants
                cell.number_format = "# ##0"
                cell.alignment = Alignment(horizontal="right")
    else:
        # Perte (Net Loss -)
        ws.append(
            ["Resultat - Net Profit (+)", 0, "Resultat - Net Loss (-)", abs(resultat)]
        )
        # Formater
        for col in [1, 2, 3, 4]:
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = result_fill
            cell.font = Font(bold=True)
            if col in [2, 4]:  # Montants
                cell.number_format = "# ##0"
                cell.alignment = Alignment(horizontal="right")

    ws.append([""] * 4)

    # TOTAUX ÉQUILIBRÉS (principe de la partie double)
    total_fill = PatternFill(
        start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"
    )

    if resultat >= 0:
        # Si profit: Charges réelles + Résultat = Produits
        total_charges_equilibre = total_charges + resultat
        total_produits_equilibre = total_produits
    else:
        # Si perte: Charges réelles = Produits + Résultat
        total_charges_equilibre = total_charges
        total_produits_equilibre = total_produits + abs(resultat)

    ws.append(
        [
            "TOTAL CHARGES",
            total_charges_equilibre,
            "TOTAL PRODUITS",
            total_produits_equilibre,
        ]
    )

    for col in range(1, 5):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        if col in [2, 4]:
            cell.number_format = "# ##0"
            cell.alignment = Alignment(horizontal="right")

    # Ajuster les largeurs de colonnes
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15

    return ws


def add_compte_resultat_sheet(wb: Workbook, cr: Dict):
    """
    Ajoute la feuille Compte de Résultat synthétique (norme SYSCOHADA)

    Format: Charges à gauche (colonnes A-C), Produits à droite (colonnes D-F)
    Structure: Exploitation, Financier, HAO
    """
    logger.info("Ajout de la feuille Compte de Résultat")

    ws = wb.create_sheet("CR SYNTH")

    # Masquer le quadrillage des cellules
    ws.sheet_view.showGridLines = False

    # Appliquer une police professionnelle pour toute la feuille
    default_font = Font(name="Calibri", size=11)

    charges = cr["charges"]
    produits = cr["produits"]

    # Ligne 1: Headers CHARGES et PRODUITS
    ws.append(["Charges", "", "", "Produits", "", ""])
    apply_header_style(ws, 1, 6)

    # Ligne 2: Achats marchandises | Ventes marchandises
    ws.append(
        [
            "Achats de marchandises, de matières & fournitures liées",
            "",
            charges["achats_marchandises"],
            "Ventes de marchandises",
            "",
            produits["ventes_marchandises"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 4: Variation stocks | Commissions
    ws.append(
        [
            "Variation de stock de mses",
            "",
            charges["variation_stocks"],
            "Commissions",
            "",
            produits["commissions"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 6: Autres achats | Autres produits
    ws.append(
        [
            "Autres achats",
            "",
            charges["autres_achats"],
            "Autres produits (Accessoires)",
            "",
            produits["autres_produits"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 8: Transports | Subventions
    ws.append(
        [
            "Transports",
            "",
            charges["transports"],
            "Subvention d'exploitation",
            "",
            produits["subventions_exploitation"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 10: Services extérieurs
    ws.append(["Services exterieurs ", "", charges["services_exterieurs"], "", "", ""])

    ws.append([])  # Ligne vide

    # Ligne 12: Impôts & taxes
    ws.append(["Impots & taxes", "", charges["impots_taxes"], "", "", ""])

    ws.append([])  # Ligne vide

    # Ligne 14: Charges du personnel
    ws.append(["Charges du personnel", "", charges["charges_personnel"], "", "", ""])

    ws.append([])  # Ligne vide

    # Ligne 16: Autres charges
    ws.append(["Autres charges ", "", charges["autres_charges"], "", "", ""])

    ws.append([])  # Ligne vide

    # Ligne 18: Frais financiers
    ws.append(["Frais financiers ", "", charges["frais_financiers"], "", "", ""])

    ws.append([])  # Ligne vide

    # Ligne 20: Dotations aux amortissements
    ws.append(
        [
            "Dotations aux amortissements",
            "",
            charges["dotations_amortissements"],
            "",
            "",
            "",
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 22: Total Exploitation (avec bordure épaisse pour séparer)
    ws.append(
        [
            "Charges d'exploitations",
            "",
            charges["total_exploitation"],
            "Produits d'exploitation",
            "",
            produits["total_exploitation"],
        ]
    )
    apply_section_separator(ws, 22, 6)  # Bordure épaisse en bas

    ws.append([])  # Ligne vide

    # SECTION FINANCIÈRE
    # Ligne 24: Intérêts payés | Revenus financiers
    ws.append(
        [
            "Interets payés",
            "",
            charges["interets_payes"],
            "Revenus financiers",
            "",
            produits["revenus_financiers"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 26: Escomptes accordés | Escomptes obtenus
    ws.append(
        [
            "Escompte accordés",
            "",
            charges["escomptes_accordes"],
            "Escomptes Obtenus",
            "",
            produits["escomptes_obtenus"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 28: Pertes de change | Gains de change
    ws.append(
        [
            "Pertes de changes",
            "",
            charges["pertes_change"],
            "Gain de change",
            "",
            produits["gains_change"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 30: Total Financier (avec bordure épaisse)
    ws.append(
        [
            "Charges Financières",
            "",
            charges["total_financier"],
            "Produits financiers",
            "",
            produits["total_financier"],
        ]
    )
    apply_section_separator(ws, 30, 6)  # Bordure épaisse en bas

    ws.append([])  # Ligne vide

    # SECTION HAO
    # Ligne 32: Valeur comptable cession | Produits cession
    ws.append(
        [
            "Valeur comptable de cession",
            "",
            0,
            "Produits cession",
            "",
            produits["produits_cession"],
        ]
    )

    ws.append([])  # Ligne vide

    # Ligne 34: Charges HAO | Produits HAO
    ws.append(
        [
            "Charges HAO",
            "",
            charges["charges_hao"],
            "Produits HAO",
            "",
            produits["total_hao"],
        ]
    )
    apply_section_separator(ws, 34, 6)  # Bordure épaisse en bas

    ws.append([])  # Ligne vide

    # SECTION RÉSULTAT
    # Ligne 36: Impôt sur les sociétés
    ws.append(["Impot sur les sociétes", "", 0, "", "", ""])

    ws.append([])  # Ligne vide

    # Ligne 38: Résultat Net
    resultat_label = "Net Profit (+)" if cr["resultat"] > 0 else "Net Loss (-)"
    ws.append(
        [
            "Resultat",
            resultat_label,
            abs(cr["resultat"]),
            "Resultat",
            resultat_label if cr["resultat"] < 0 else "",
            abs(cr["resultat"]) if cr["resultat"] < 0 else "",
        ]
    )
    apply_result_style(ws, 38, 6)  # Style distinctif pour le résultat

    ws.append([])  # Ligne vide

    # Ligne 40: Totaux finaux
    ws.append(
        [
            "TOTAL CHARGES",
            "",
            cr["total_charges"],
            "TOTAL PRODUITS",
            "",
            cr["total_produits"],
        ]
    )
    apply_total_style(ws, 40, 6)

    ws.append([])  # Ligne vide

    # Ligne 42: Écart (vérification)
    ws.append(["Ecart", "", f"=C40-F40", "", "", ""])

    # Appliquer le formatage des nombres
    for row in [
        2,
        4,
        6,
        8,
        10,
        12,
        14,
        16,
        18,
        20,
        22,
        24,
        26,
        28,
        30,
        32,
        34,
        36,
        38,
        40,
        42,
    ]:
        apply_number_format(ws, "C", row, row)  # Colonne Charges
        apply_number_format(ws, "F", row, row)  # Colonne Produits

    # Appliquer la police Calibri 11 à toutes les cellules
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            if (
                cell.font.name != "Calibri" or not cell.font.bold
            ):  # Ne pas modifier les headers
                cell.font = default_font

    # Ajuster la largeur des colonnes
    adjust_column_widths(ws)

    # Garantir une largeur minimale pour les colonnes de montants
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width, 18)
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width, 18)


def add_sig_sheet(wb: Workbook, sig: Dict):
    """
    Ajoute la feuille SIG (Soldes Intermédiaires de Gestion) - Format SYSCOHADA

    Le SIG présente la formation du résultat par étapes successives selon OHADA.
    Chaque ligne a un code de référence (TA, RA, XA, etc.)
    """
    logger.info("Ajout de la feuille SIG")

    ws = wb.create_sheet("SIG")

    # Ligne 1: Titre
    ws.append(["Soldes intermédiaires de gestion (SIG)", "", "", "", "", ""])
    apply_header_style(ws, 1, 6)

    ws.append([])  # Ligne vide

    # Ligne 3: En-tête
    ws.append(["#NAME?", "LIBELLES", "", "", "EXERCICE AU 30/09/2025", "EXERCICE "])
    # Ligne 4: Sous-en-tête
    ws.append(["", "", "", "", "NET", "NET"])
    apply_header_style(ws, 3, 6)
    apply_header_style(ws, 4, 6)

    # I. ACTIVITÉ D'EXPLOITATION
    # Ligne 5: TA - Ventes de marchandises
    ws.append(["TA", "Ventes de marchandises", "A", "+", sig["ventes_marchandises"], 0])
    # Ligne 6: RA - Achats de marchandises
    ws.append(["RA", "Achats de marchandises", "", "-", sig["achats_marchandises"], 0])
    # Ligne 7: RB - Variation stocks
    ws.append(["RB", "Variation de stocks de marchandises", "", "-/+", 0, 0])
    # Ligne 8: XA - Marge commerciale
    ws.append(
        [
            "XA",
            "MARGE COMMERCIALE  (somme TA à RB)",
            "",
            "",
            sig["marge_commerciale"],
            0,
        ]
    )

    # Ligne 9: TB - Ventes produits fabriqués
    ws.append(["TB", "Ventes de produits fabriqués", "B", "+", 0, 0])
    # Ligne 10: TC - Travaux et services
    ws.append(["TC", "Travaux, services vendus", "C", "+", 0, 0])
    # Ligne 11: TD - Commissions
    ws.append(["TD", "Commissions & courtages", "D", "+", sig["production_vendue"], 0])
    # Ligne 12: XB - Chiffre d'affaires
    ws.append(
        ["XB", "CHIFFRE D'AFFAIRES (A+B+C+D)", "", "", sig["chiffre_affaires"], 0]
    )

    # Ligne 13: TE - Production stockée
    ws.append(["TE", "Production stockée", "", "-/+", 0, 0])
    # Ligne 14: TG - Subventions
    ws.append(["TG", "Subventions d'exploitation", "", "+", 0, 0])
    # Ligne 15: TH - Autres produits
    ws.append(["TH", "Autres produits", "", "+", 0, 0])
    # Ligne 16: RC - Achats matières
    ws.append(
        ["RC", "Achats de matieres premières et fournitures liées", "", "-", 0, 0]
    )
    # Ligne 17: RD - Variation stocks matières
    ws.append(
        [
            "RD",
            "Variation de stock de matières premieres & fournitures liées",
            "",
            "-/+",
            0,
            0,
        ]
    )
    # Ligne 18: RE - Autres achats
    ws.append(["RE", "Autres achats ", "", "-", 0, 0])
    # Ligne 19: RF - Variation autres stocks
    ws.append(["RF", "Variation de stocks d'autres approvisionements", "", "-/+", 0, 0])
    # Ligne 20: RG - Transports
    ws.append(["RG", "Transports", "", "-", 0, 0])
    # Ligne 21: RH - Services extérieurs
    ws.append(["RH", "Services exterieurs", "", "-", -sig["services_exterieurs"], 0])
    # Ligne 22: RI - Impôts et taxes
    ws.append(["RI", "Impots et taxes", "", "-", -sig["impots_taxes"], 0])
    # Ligne 23: RJ - Autres charges
    ws.append(["RJ", "Autres charges", "", "-", -sig["autres_charges"], 0])
    # Ligne 24: XC - Valeur ajoutée
    ws.append(
        [
            "XC",
            "VALEUR AJOUTÉE (XB+RA+RB) +(somme TE à RI)",
            "",
            "",
            sig["valeur_ajoutee"],
            0,
        ]
    )

    # Ligne 25: RK - Charges personnel
    ws.append(["RK", "Charges du personnel ", "", "-", -sig["charges_personnel"], 0])
    # Ligne 26: XD - EBE
    ws.append(["XD", "EXCÉDENT BRUT D'EXPLOITATION (XC+RK)", "", "", sig["ebe"], 0])

    # Ligne 27: TJ - Reprises
    ws.append(
        [
            "TJ",
            "Reprises d'amortissments, de provisions et dépréciations",
            "",
            "+",
            0,
            0,
        ]
    )
    # Ligne 28: RL - Dotations
    ws.append(
        [
            "RL",
            "Dotations aux amortissements, aux provisions et dépréciations",
            "",
            "-",
            -sig["dotations_amortissements"],
            0,
        ]
    )
    # Ligne 29: XE - Résultat d'exploitation
    ws.append(
        [
            "XE",
            "RÉSULTAT D'EXPLOITATION (XD+TJ+RL)",
            "",
            "",
            sig["resultat_exploitation"],
            0,
        ]
    )

    # II. ACTIVITÉ FINANCIÈRE
    # Ligne 30: TK - Revenus financiers
    ws.append(["TK", "Revenus financiers et assimilés", "", "+", 0, 0])
    # Ligne 31: TL - Reprises financières
    ws.append(
        ["TL", "Reprises de provisions et dépréciations financieres ", "", "+", 0, 0]
    )
    # Ligne 32: TM - Transferts charges
    ws.append(["TM", "Transferts de charges financieres ", "", "+", 0, 0])
    # Ligne 33: RM - Frais financiers
    ws.append(
        [
            "RM",
            "Frais finaciers et charges Assimilés ",
            "",
            "-",
            -sig["frais_financiers"],
            0,
        ]
    )
    # Ligne 34: RN - Dotations financières
    ws.append(
        [
            "RN",
            "Dotations aux provisions et aux dépréciations finacières",
            "",
            "-",
            0,
            0,
        ]
    )
    # Ligne 35: XF - Résultat financier
    resultat_financier = -sig["frais_financiers"]
    ws.append(
        ["XF", "RÉSULTAT FINANCIER ( somme TK à RN)", "", "", resultat_financier, 0]
    )

    # Ligne 36: XG - Résultat activités ordinaires
    resultat_ordinaires = sig["resultat_exploitation"] + resultat_financier
    ws.append(
        [
            "XG",
            "RESULTAT DES ACTIVITES ORDINAIRES (XE+XF)",
            "",
            "",
            resultat_ordinaires,
            0,
        ]
    )

    # III. HORS ACTIVITÉS ORDINAIRES (HAO)
    # Ligne 37: TN - Produits cessions
    ws.append(["TN", "Produits des cessions d'immobilisations", "", "+", 0, 0])
    # Ligne 38: TO - Autres produits HAO
    ws.append(["TO", "Autres produits HAO", "", "+", 0, 0])
    # Ligne 39: RO - Valeurs comptables cessions
    ws.append(
        ["RO", "Valeurs comptables des cessions d'immobilisations", "", "-", 0, 0]
    )
    # Ligne 40: RP - Autres charges HAO
    ws.append(["RP", "Autres charges HAO", "", "-", 0, 0])
    # Ligne 41: XH - Résultat HAO
    ws.append(
        ["XH", "RESULTAT HORS  ACTIVITES ORDINAIRES (somme TN à RP)", "", "", 0, 0]
    )

    # Ligne 42: RQ - Participation
    ws.append(["RQ", "Participation des travailleurs", "", "-", "", ""])
    # Ligne 43: RS - Impôts sur résultat
    ws.append(["RS", "Impots sur le résultat", "", "-", "", 0])
    # Ligne 44: XI - Résultat net
    ws.append(
        [
            "XI",
            "RÉSULTAT NET DE L'EXERCICE (XG+XH+RQ+RS)",
            "",
            "",
            sig["resultat_net"],
            0,
        ]
    )

    # Appliquer le formatage des nombres sur la colonne E (Exercice actuel)
    for row in range(5, 45):
        apply_number_format(ws, "E", row, row)

    # Appliquer un style pour les lignes de totaux (XA, XB, XC, XD, XE, XF, XG, XH, XI)
    total_rows = [8, 12, 24, 26, 29, 35, 36, 41, 44]
    for row in total_rows:
        apply_total_style(ws, row, 6)

    # Ajuster la largeur des colonnes
    adjust_column_widths(ws)


def detect_available_months(df: pd.DataFrame) -> list:
    """
    Détecte les mois présents dans le Grand Livre

    Args:
        df: DataFrame du Grand Livre

    Returns:
        Liste des noms de mois présents (ex: ['Janvier', 'Février', ..., 'Octobre'])
    """
    # Dictionnaire de conversion
    mois_noms = {
        1: "Janvier",
        2: "Février",
        3: "Mars",
        4: "Avril",
        5: "Mai",
        6: "Juin",
        7: "Juillet",
        8: "Août",
        9: "Septembre",
        10: "Octobre",
        11: "Novembre",
        12: "Décembre",
    }

    # Filtrer les lignes avec date valide
    df_with_date = df[df["date"].notna()]

    if len(df_with_date) == 0:
        # Par défaut, retourner Janvier à Septembre
        return [mois_noms[i] for i in range(1, 10)]

    # Extraire les numéros de mois présents
    mois_present = sorted(df_with_date["date"].dt.month.unique())

    # Convertir en noms de mois
    mois_list = [mois_noms[m] for m in mois_present if m in mois_noms]

    logger.info(f"Mois détectés dans les données: {', '.join(mois_list)}")

    return mois_list


def add_suivi_activite_sheet(wb: Workbook, df: pd.DataFrame, client_code: str = None):
    """
    Ajoute la feuille SUIVI ACTIVITE - Tableau de bord budgétaire mensuel

    Ce tableau présente l'évolution mensuelle des charges et produits
    par catégories budgétaires pour le pilotage de l'activité.

    Args:
        wb: Workbook Excel
        df: DataFrame du Grand Livre
        client_code: Code du client pour utiliser son mapping spécifique (optionnel)
    """
    if client_code:
        logger.info(
            f"Ajout de la feuille Suivi d'Activité pour le client: {client_code}"
        )
    else:
        logger.info("Ajout de la feuille Suivi d'Activité (mapping par défaut)")

    from modules.data_processor import (
        prepare_suivi_activite_detaille,
        get_suivi_activite_categories,
    )

    # Préparer les données mensuelles DÉTAILLÉES par compte avec le mapping client si fourni
    suivi_data_detaille = prepare_suivi_activite_detaille(df, client_code=client_code)

    # Récupérer les catégories organisées par type (dynamique selon le client)
    categories = get_suivi_activite_categories(client_code=client_code)

    ws = wb.create_sheet("SUIVI ACTIVITE")

    # Extraire l'année des données
    df_with_date = df[df["date"].notna()]
    if len(df_with_date) > 0:
        annee = df_with_date["date"].dt.year.mode()[0]
    else:
        annee = 2025

    # Détecter les mois présents dans les données
    mois_disponibles = detect_available_months(df)
    nb_mois = len(mois_disponibles)

    # Calculer le nombre total de colonnes
    # 4 colonnes fixes (A, B, C, D) + nb_mois + 4 colonnes d'analyse (TOTAL, Écart, Variation, Variation %)
    nb_colonnes_total = 4 + nb_mois + 4

    # Position de la première colonne de mois (E = 5)
    col_premier_mois = 5
    # Position de la dernière colonne de mois
    col_dernier_mois = col_premier_mois + nb_mois - 1
    # Position de TOTAL ANNEE
    col_total_annee = col_dernier_mois + 1
    # Position des colonnes d'analyse
    col_ecart_budget = col_total_annee + 1
    col_variation = col_ecart_budget + 1
    col_variation_pct = col_variation + 1

    logger.info(
        f"Configuration dynamique: {nb_mois} mois détectés, {nb_colonnes_total} colonnes totales"
    )

    # Ligne 1: Titre principal centré et fusionné
    ws.append(["", f"Tableau de suivi BAMBOO IMMO {annee}"])
    # Fusionner sur toutes les colonnes pour un titre large
    end_col_letter = get_column_letter(nb_colonnes_total)
    ws.merge_cells(f"A1:{end_col_letter}1")

    # Style spécial pour le titre principal
    title_cell = ws["A1"]
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.font = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30  # Hauteur augmentée pour le titre

    # Lignes vides
    for _ in range(3):
        ws.append([""] * nb_colonnes_total)

    # Ligne 5: En-têtes avec les mois + TOTAL ANNEE + colonnes d'analyse
    mois_headers = [
        "",
        "",
        "Rappel année N-1",
        "BUDGET PREVI",
    ]
    # Ajouter les mois détectés
    mois_headers.extend(mois_disponibles)
    # Ajouter les colonnes d'analyse
    mois_headers.extend(
        [
            "TOTAL ANNEE",
            "Ecart au budget",
            "Variation",
            "Variation %",
        ]
    )

    ws.append(mois_headers)
    apply_header_style(ws, 5, nb_colonnes_total)

    # SECTION 1: CHARGES DE PERSONNEL (dynamique selon le mapping client)
    personnel_categories = categories["personnel"]

    if personnel_categories:
        personnel_start_row = 6
        personnel_subtotal_rows = []  # Traquer les lignes de sous-totaux uniquement
        detail_row_counter = 0  # Compteur pour alternance de couleur

        for cat_key, cat_label in personnel_categories:
            # Vérifier si cette catégorie a des comptes
            cat_data = suivi_data_detaille.get(cat_key, {})
            comptes = cat_data.get("comptes", {})

            if comptes:
                # Ajouter une ligne d'en-tête de catégorie avec indentation
                category_row = ["", f"  {cat_label}", "", ""]  # Indentation 1 niveau
                for mois in mois_disponibles:
                    category_row.append("")
                # Ajouter colonnes d'analyse vides
                category_row.extend(["", "", "", ""])
                ws.append(category_row)
                apply_category_header_style(ws, ws.max_row, nb_colonnes_total)

                # Pour chaque compte de cette catégorie, afficher une ligne détaillée
                for compte_num, compte_info in comptes.items():
                    # Format: "    Libellé du compte" (double indentation)
                    compte_label = f"    {compte_info['libelle'][:50]}"
                    row_data = ["", compte_label, 0, 0]  # N-1 et Budget à 0
                    detail_row_counter += 1

                    # Ajouter les montants mensuels pour ce compte (dynamique)
                    for mois in mois_disponibles:
                        montant = compte_info["mois_data"].get(mois, 0)
                        # Pour les charges, inverser le signe
                        montant = -montant if montant > 0 else montant
                        row_data.append(montant if montant != 0 else 0)
                    ws.append(row_data)
                    # Ajouter les formules pour les colonnes calculées (positions dynamiques)
                    current_row = ws.max_row
                    # TOTAL ANNEE
                    total_cell = ws.cell(row=current_row, column=col_total_annee)
                    col_premier_mois_letter = get_column_letter(col_premier_mois)
                    col_dernier_mois_letter = get_column_letter(col_dernier_mois)
                    total_cell.value = f"=SUM({col_premier_mois_letter}{current_row}:{col_dernier_mois_letter}{current_row})"
                    total_cell.number_format = "# ##0"
                    # Écart au budget = TOTAL ANNEE - Budget
                    ecart_cell = ws.cell(row=current_row, column=col_ecart_budget)
                    col_total_letter = get_column_letter(col_total_annee)
                    ecart_cell.value = (
                        f"={col_total_letter}{current_row}-D{current_row}"
                    )
                    ecart_cell.number_format = "# ##0"
                    # Variation = TOTAL ANNEE - Rappel N-1
                    variation_cell = ws.cell(row=current_row, column=col_variation)
                    variation_cell.value = (
                        f"={col_total_letter}{current_row}-C{current_row}"
                    )
                    variation_cell.number_format = "# ##0"
                    # Variation % = (TOTAL ANNEE - Rappel N-1) / Rappel N-1 * 100
                    variation_pct_cell = ws.cell(
                        row=current_row, column=col_variation_pct
                    )
                    variation_pct_cell.value = f"=IF(C{current_row}=0,0,({col_total_letter}{current_row}-C{current_row})/C{current_row}*100)"
                    variation_pct_cell.number_format = "0.00%"
                    # Appliquer le style de ligne de détail avec alternance
                    apply_detail_row_style(ws, current_row, nb_colonnes_total, is_odd=(detail_row_counter % 2 == 1))

                # Ajouter la ligne de sous-total pour cette catégorie
                row_data = ["", f"  Total {cat_label}", 0, 0]  # Indentation 1 niveau
                for mois in mois_disponibles:
                    montant = cat_data["total"].get(mois, 0)
                    row_data.append(montant if montant != 0 else 0)
                ws.append(row_data)
                # Traquer cette ligne de sous-total
                personnel_subtotal_rows.append(ws.max_row)
                # Ajouter les formules pour les colonnes calculées (positions dynamiques)
                current_row = ws.max_row
                # TOTAL ANNEE
                total_cell = ws.cell(row=current_row, column=col_total_annee)
                col_premier_mois_letter = get_column_letter(col_premier_mois)
                col_dernier_mois_letter = get_column_letter(col_dernier_mois)
                total_cell.value = f"=SUM({col_premier_mois_letter}{current_row}:{col_dernier_mois_letter}{current_row})"
                total_cell.number_format = "# ##0"
                # Écart au budget
                ecart_cell = ws.cell(row=current_row, column=col_ecart_budget)
                col_total_letter = get_column_letter(col_total_annee)
                ecart_cell.value = f"={col_total_letter}{current_row}-D{current_row}"
                ecart_cell.number_format = "# ##0"
                # Variation
                variation_cell = ws.cell(row=current_row, column=col_variation)
                variation_cell.value = (
                    f"={col_total_letter}{current_row}-C{current_row}"
                )
                variation_cell.number_format = "# ##0"
                # Variation %
                variation_pct_cell = ws.cell(row=current_row, column=col_variation_pct)
                variation_pct_cell.value = f"=IF(C{current_row}=0,0,({col_total_letter}{current_row}-C{current_row})/C{current_row}*100)"
                variation_pct_cell.number_format = "0.00%"
                # Appliquer le style de sous-total
                apply_subtotal_style(ws, ws.max_row, nb_colonnes_total)
            else:
                # Si pas de comptes, afficher juste la catégorie avec des zéros
                row_data = ["", cat_label, 0, 0] + [0] * (
                    nb_mois + 4
                )  # mois + 4 colonnes d'analyse
                ws.append(row_data)
                personnel_subtotal_rows.append(ws.max_row)

        # Ligne Total charges de personnel - sommer UNIQUEMENT les lignes de sous-totaux
        total_personnel_row_num = ws.max_row + 1
        total_personnel_row = [
            "",
            "Total charges de personnel",
            0,  # N-1 (pas utilisé)
            0,  # Budget (pas utilisé)
        ]

        # Pour chaque colonne de mois + TOTAL ANNEE + Colonnes d'analyse
        for col_idx in range(col_premier_mois, col_variation_pct + 1):  # Dynamique
            col_letter = get_column_letter(col_idx)
            # Sommer uniquement les lignes de sous-totaux
            if personnel_subtotal_rows:
                formula_parts = [
                    f"{col_letter}{row}" for row in personnel_subtotal_rows
                ]
                total_personnel_row.append(f"={'+'.join(formula_parts)}")
            else:
                total_personnel_row.append(0)

        ws.append(total_personnel_row)
        apply_total_style(ws, ws.max_row, nb_colonnes_total)
    else:
        # Si pas de catégories personnel, créer quand même une ligne totale à 0
        personnel_start_row = 6
        total_personnel_row_num = 6
        ws.append(["", "Total charges de personnel", 0, 0] + [0] * (nb_mois + 4))
        apply_total_style(ws, ws.max_row, nb_colonnes_total)

    # SECTION 2: AUTRES CHARGES (dynamique selon le mapping client)
    autres_charges_categories = categories["charges"]

    if autres_charges_categories:
        autres_charges_subtotal_rows = []  # Traquer les lignes de sous-totaux uniquement

        for cat_key, cat_label in autres_charges_categories:
            # Vérifier si cette catégorie a des comptes
            cat_data = suivi_data_detaille.get(cat_key, {})
            comptes = cat_data.get("comptes", {})

            if comptes:
                # Pour chaque compte de cette catégorie, afficher une ligne détaillée
                for compte_num, compte_info in comptes.items():
                    # Format: "  Libellé du compte" (sans le numéro de compte)
                    compte_label = f"  {compte_info['libelle'][:50]}"
                    row_data = ["", compte_label, 0, 0]  # N-1 et Budget à 0

                    # Ajouter les montants mensuels pour ce compte (dynamique)
                    for mois in mois_disponibles:
                        montant = compte_info["mois_data"].get(mois, 0)
                        # Pour les charges, inverser le signe
                        montant = -montant if montant > 0 else montant
                        row_data.append(montant if montant != 0 else 0)
                    ws.append(row_data)
                    # Ajouter les formules pour les colonnes calculées (positions dynamiques)
                    current_row = ws.max_row
                    # TOTAL ANNEE
                    total_cell = ws.cell(row=current_row, column=col_total_annee)
                    col_premier_mois_letter = get_column_letter(col_premier_mois)
                    col_dernier_mois_letter = get_column_letter(col_dernier_mois)
                    total_cell.value = f"=SUM({col_premier_mois_letter}{current_row}:{col_dernier_mois_letter}{current_row})"
                    total_cell.number_format = "# ##0"
                    # Écart au budget = TOTAL ANNEE - Budget
                    ecart_cell = ws.cell(row=current_row, column=col_ecart_budget)
                    col_total_letter = get_column_letter(col_total_annee)
                    ecart_cell.value = (
                        f"={col_total_letter}{current_row}-D{current_row}"
                    )
                    ecart_cell.number_format = "# ##0"
                    # Variation = TOTAL ANNEE - Rappel N-1
                    variation_cell = ws.cell(row=current_row, column=col_variation)
                    variation_cell.value = (
                        f"={col_total_letter}{current_row}-C{current_row}"
                    )
                    variation_cell.number_format = "# ##0"
                    # Variation % = (TOTAL ANNEE - Rappel N-1) / Rappel N-1 * 100
                    variation_pct_cell = ws.cell(
                        row=current_row, column=col_variation_pct
                    )
                    variation_pct_cell.value = f"=IF(C{current_row}=0,0,({col_total_letter}{current_row}-C{current_row})/C{current_row}*100)"
                    variation_pct_cell.number_format = "0.00%"

                # Ajouter la ligne de sous-total pour cette catégorie
                row_data = ["", f"Total {cat_label}", 0, 0]
                for mois in mois_disponibles:
                    montant = cat_data["total"].get(mois, 0)
                    row_data.append(montant if montant != 0 else 0)
                ws.append(row_data)
                # Traquer cette ligne de sous-total
                autres_charges_subtotal_rows.append(ws.max_row)
                # Ajouter la formule TOTAL ANNEE pour le sous-total (DYNAMIQUE)
                current_row = ws.max_row
                total_cell = ws.cell(row=current_row, column=col_total_annee)
                col_premier_mois_letter = get_column_letter(col_premier_mois)
                col_dernier_mois_letter = get_column_letter(col_dernier_mois)
                total_cell.value = f"=SUM({col_premier_mois_letter}{current_row}:{col_dernier_mois_letter}{current_row})"
                total_cell.number_format = "# ##0"
                # Mettre en gras la ligne de sous-total
                for cell in ws[ws.max_row]:
                    if cell.column <= nb_colonnes_total:
                        cell.font = Font(bold=True)
            else:
                # Si pas de comptes, afficher juste la catégorie avec des zéros
                row_data = ["", cat_label, 0, 0] + [0] * (
                    nb_mois + 4
                )  # mois + 4 colonnes d'analyse
                ws.append(row_data)
                autres_charges_subtotal_rows.append(ws.max_row)

        # Ligne Total autres charges - sommer UNIQUEMENT les lignes de sous-totaux
        total_autres_charges_row_num = ws.max_row + 1
        total_autres_row = [
            "",
            "Autres charges",
            0,  # N-1 (pas utilisé)
            0,  # Budget (pas utilisé)
        ]

        # Pour chaque colonne de mois + TOTAL ANNEE
        for col_idx in range(5, 18):  # Janvier à Variation % (colonnes E à Q)
            col_letter = get_column_letter(col_idx)
            # Sommer uniquement les lignes de sous-totaux
            if autres_charges_subtotal_rows:
                formula_parts = [
                    f"{col_letter}{row}" for row in autres_charges_subtotal_rows
                ]
                total_autres_row.append(f"={'+'.join(formula_parts)}")
            else:
                total_autres_row.append(0)

        ws.append(total_autres_row)
        apply_total_style(ws, ws.max_row, nb_colonnes_total)
    else:
        # Si pas de catégories charges, créer quand même une ligne totale à 0
        total_autres_charges_row_num = ws.max_row + 1
        ws.append(["", "Autres charges", 0, 0] + [0] * (nb_mois + 4))
        apply_total_style(ws, ws.max_row, nb_colonnes_total)

    # Lignes vides
    for _ in range(3):
        ws.append([""] * nb_colonnes_total)

    # Ligne Total charges (Personnel + Autres)
    total_charges_row_num = ws.max_row + 1
    total_charges_row = [
        "",
        "Total charges",
        f"=+C{total_personnel_row_num}+C{total_autres_charges_row_num}",
        f"=+D{total_personnel_row_num}+D{total_autres_charges_row_num}",
    ]
    for col_idx in range(col_premier_mois, col_variation_pct + 1):
        col_letter = get_column_letter(col_idx)
        total_charges_row.append(
            f"=+{col_letter}{total_personnel_row_num}+{col_letter}{total_autres_charges_row_num}"
        )
    # Ajouter la colonne TOTAL ANNEE (colonne N = 14)
    total_charges_row.append(
        f"=+N{total_personnel_row_num}+N{total_autres_charges_row_num}"
    )
    ws.append(total_charges_row)
    apply_total_style(ws, ws.max_row, nb_colonnes_total)

    # Ajouter une bordure supérieure épaisse pour TOTAL CHARGES
    for col in range(1, nb_colonnes_total + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(
            top=Side(style="medium", color="000000"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

    # Lignes vides
    for _ in range(2):
        ws.append([""] * nb_colonnes_total)

    # SECTION 3: PRODUITS (dynamique selon le mapping client)
    produits_categories = categories["produits"]

    if produits_categories:
        produits_subtotal_rows = []  # Traquer les lignes de sous-totaux uniquement

        for cat_key, cat_label in produits_categories:
            # Vérifier si cette catégorie a des comptes
            cat_data = suivi_data_detaille.get(cat_key, {})
            comptes = cat_data.get("comptes", {})

            if comptes:
                # Pour chaque compte de cette catégorie, afficher une ligne détaillée
                for compte_num, compte_info in comptes.items():
                    # Format: "  Libellé du compte" (sans le numéro de compte)
                    compte_label = f"  {compte_info['libelle'][:50]}"
                    row_data = ["", compte_label, 0, 0]  # N-1 et Budget à 0

                    # Ajouter les montants mensuels pour ce compte
                    for mois in mois_disponibles:
                        montant = compte_info["mois_data"].get(mois, 0)
                        # Pour les produits, garder le signe positif
                        row_data.append(montant if montant != 0 else 0)
                    ws.append(row_data)
                    # Ajouter les formules pour les colonnes calculées (positions dynamiques)
                    current_row = ws.max_row
                    # TOTAL ANNEE
                    total_cell = ws.cell(row=current_row, column=col_total_annee)
                    col_premier_mois_letter = get_column_letter(col_premier_mois)
                    col_dernier_mois_letter = get_column_letter(col_dernier_mois)
                    total_cell.value = f"=SUM({col_premier_mois_letter}{current_row}:{col_dernier_mois_letter}{current_row})"
                    total_cell.number_format = "# ##0"
                    # Écart au budget = TOTAL ANNEE - Budget
                    ecart_cell = ws.cell(row=current_row, column=col_ecart_budget)
                    col_total_letter = get_column_letter(col_total_annee)
                    ecart_cell.value = (
                        f"={col_total_letter}{current_row}-D{current_row}"
                    )
                    ecart_cell.number_format = "# ##0"
                    # Variation = TOTAL ANNEE - Rappel N-1
                    variation_cell = ws.cell(row=current_row, column=col_variation)
                    variation_cell.value = (
                        f"={col_total_letter}{current_row}-C{current_row}"
                    )
                    variation_cell.number_format = "# ##0"
                    # Variation % = (TOTAL ANNEE - Rappel N-1) / Rappel N-1 * 100
                    variation_pct_cell = ws.cell(
                        row=current_row, column=col_variation_pct
                    )
                    variation_pct_cell.value = f"=IF(C{current_row}=0,0,({col_total_letter}{current_row}-C{current_row})/C{current_row}*100)"
                    variation_pct_cell.number_format = "0.00%"

                # Ajouter la ligne de sous-total pour cette catégorie
                row_data = ["", f"Total {cat_label}", 0, 0]
                for mois in mois_disponibles:
                    montant = cat_data["total"].get(mois, 0)
                    row_data.append(montant if montant != 0 else 0)
                ws.append(row_data)
                # Traquer cette ligne de sous-total
                produits_subtotal_rows.append(ws.max_row)
                # Ajouter la formule TOTAL ANNEE pour le sous-total (DYNAMIQUE)
                current_row = ws.max_row
                total_cell = ws.cell(row=current_row, column=col_total_annee)
                col_premier_mois_letter = get_column_letter(col_premier_mois)
                col_dernier_mois_letter = get_column_letter(col_dernier_mois)
                total_cell.value = f"=SUM({col_premier_mois_letter}{current_row}:{col_dernier_mois_letter}{current_row})"
                total_cell.number_format = "# ##0"
                # Mettre en gras la ligne de sous-total
                for cell in ws[ws.max_row]:
                    if cell.column <= nb_colonnes_total:
                        cell.font = Font(bold=True)
            else:
                # Si pas de comptes, afficher juste la catégorie avec des zéros
                row_data = ["", cat_label, 0, 0] + [0] * (
                    nb_mois + 4
                )  # mois + 4 colonnes d'analyse
                ws.append(row_data)
                produits_subtotal_rows.append(ws.max_row)

        # Ligne vide de séparation
        ws.append([""] * nb_colonnes_total)

        # Ligne Total CA - sommer UNIQUEMENT les lignes de sous-totaux
        total_ca_row_num = ws.max_row + 1
        total_ca_row = [
            "",
            "CA",  # Libellé simplifié
            0,  # N-1 (pas utilisé)
            0,  # Budget (pas utilisé)
        ]

        # Pour chaque colonne de mois + colonnes d'analyse
        for col_idx in range(col_premier_mois, col_variation_pct + 1):
            col_letter = get_column_letter(col_idx)
            # Sommer uniquement les lignes de sous-totaux
            if produits_subtotal_rows:
                formula_parts = [f"{col_letter}{row}" for row in produits_subtotal_rows]
                total_ca_row.append(f"={'+'.join(formula_parts)}")
            else:
                total_ca_row.append(0)

        ws.append(total_ca_row)
        apply_total_style(ws, ws.max_row, nb_colonnes_total)
    else:
        # Si pas de catégories produits, créer quand même une ligne totale à 0
        total_ca_row_num = ws.max_row + 1
        ws.append([""] * nb_colonnes_total)  # Ligne vide
        ws.append(["", "CA", 0, 0] + [0] * (nb_mois + 4))
        apply_total_style(ws, ws.max_row, nb_colonnes_total)

    # Ligne vide avant résultat
    ws.append([""] * nb_colonnes_total)

    # RÉSULTAT AVANT IMPÔTS (CA + Total charges, car les charges sont déjà négatives)
    resultat_avant_row_num = ws.max_row + 1
    resultat_row = [
        "",
        "Résultat avant impôts",
        f"=C{total_ca_row_num}+C{total_charges_row_num}",  # Produits + Charges (déjà négatives)
        f"=D{total_ca_row_num}+D{total_charges_row_num}",
    ]
    # Pour toutes les colonnes de mois + analyse
    for col_idx in range(col_premier_mois, col_variation_pct + 1):
        col_letter = get_column_letter(col_idx)
        resultat_row.append(
            f"={col_letter}{total_ca_row_num}+{col_letter}{total_charges_row_num}"
        )

    ws.append(resultat_row)
    apply_result_style(ws, ws.max_row, nb_colonnes_total)  # Style vert

    # Lignes vides finales
    for _ in range(3):
        ws.append([""] * nb_colonnes_total)

    # Appliquer le formatage des nombres sur toutes les colonnes de montants
    for row in range(6, ws.max_row + 1):
        for col_idx in range(3, nb_colonnes_total + 1):  # Colonnes C à la dernière
            apply_number_format(ws, get_column_letter(col_idx), row, row)

    # Appliquer la couleur jaune aux 4 colonnes d'analyse (positions dynamiques)
    for col_num in [
        col_total_annee,
        col_ecart_budget,
        col_variation,
        col_variation_pct,
    ]:
        apply_yellow_column(ws, col_num, 5, ws.max_row)

    # Ajuster la largeur des colonnes dynamiquement
    adjust_column_widths_suivi(
        ws,
        nb_mois,
        col_premier_mois,
        col_total_annee,
        col_ecart_budget,
        col_variation,
        col_variation_pct,
    )

    # Figer les volets sur la ligne d'en-tête (ligne 5) et colonne B
    # Cela permet de garder l'en-tête visible lors du scroll
    ws.freeze_panes = "B6"  # Figer après la colonne A et la ligne 5


def apply_header_style(ws, row: int, num_cols: int):
    """Applique le style aux en-têtes - Bleu marine professionnel"""
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )


def apply_total_style(ws, row: int, num_cols: int):
    """Applique le style aux lignes de total - Orange clair professionnel"""
    total_fill = PatternFill(
        start_color="FDB462", end_color="FDB462", fill_type="solid"
    )
    total_font = Font(bold=True, size=12)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = total_fill
        cell.font = total_font
        # Bordure pour séparer
        cell.border = Border(
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )


def apply_result_style(ws, row: int, num_cols: int):
    """Applique le style au résultat - Vert clair professionnel"""
    result_fill = PatternFill(
        start_color="B3DE69", end_color="B3DE69", fill_type="solid"
    )
    result_font = Font(bold=True, size=13, name="Calibri", underline="single")

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = result_fill
        cell.font = result_font
        # Bordures renforcées pour le résultat
        cell.border = Border(
            top=Side(style="medium", color="000000"),
            bottom=Side(style="medium", color="000000"),
        )


def apply_section_separator(ws, row: int, num_cols: int):
    """Applique une bordure épaisse en bas de ligne pour séparer les sections"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        # Ajouter une bordure épaisse en bas pour séparer visuellement les sections
        cell.border = Border(
            left=cell.border.left if cell.border else None,
            right=cell.border.right if cell.border else None,
            top=cell.border.top if cell.border else None,
            bottom=Side(style="medium", color="366092"),
        )
        # Police en gras pour les totaux de section
        if not cell.font.bold:
            cell.font = Font(name="Calibri", size=11, bold=True)


def apply_category_header_style(ws, row: int, num_cols: int):
    """
    Style pour les en-têtes de catégorie (Personnel, Charges externes, etc.)
    Fond gris clair + gras + taille 10pt
    """
    category_fill = PatternFill(
        start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
    )
    category_font = Font(bold=True, size=10, name="Calibri", color="000000")

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = category_fill
        cell.font = category_font
        cell.border = Border(
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )


def apply_detail_row_style(ws, row: int, num_cols: int, is_odd: bool = False):
    """
    Style pour les lignes de détail (comptes individuels)
    Police normale + indentation + alternance de couleur optionnelle
    """
    # Alternance: gris très clair pour lignes impaires, blanc pour paires
    if is_odd:
        detail_fill = PatternFill(
            start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
        )
    else:
        detail_fill = PatternFill(fill_type=None)  # Blanc/transparent

    detail_font = Font(size=9, name="Calibri", color="404040")

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        if col == 2:  # Colonne libellé
            cell.fill = detail_fill
            cell.font = detail_font
            # Ajouter bordure pointillée légère
            cell.border = Border(
                bottom=Side(style="dotted", color="E0E0E0"),
            )
        else:
            cell.fill = detail_fill
            cell.font = Font(size=9, name="Calibri", color="202020")


def apply_subtotal_style(ws, row: int, num_cols: int):
    """
    Style pour les sous-totaux (Total Personnel, Total Charges externes, etc.)
    Fond orange clair + gras + bordure en haut
    """
    subtotal_fill = PatternFill(
        start_color="FED8B1", end_color="FED8B1", fill_type="solid"
    )
    subtotal_font = Font(bold=True, size=10, name="Calibri", color="000000")

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = subtotal_fill
        cell.font = subtotal_font
        cell.border = Border(
            top=Side(style="thin", color="CC9966"),
            bottom=Side(style="thin", color="CCCCCC"),
        )


def apply_number_format(ws, column: str, start_row: int, end_row: int):
    """Applique le format numérique aux cellules"""
    for row in range(start_row, end_row + 1):
        cell = ws[f"{column}{row}"]
        cell.number_format = "# ##0"


def apply_yellow_column(ws, column_num: int, start_row: int, end_row: int):
    """
    Applique un fond gris clair professionnel aux colonnes d'analyse
    avec mise en forme conditionnelle pour les variations
    """
    # Fond gris clair pour les colonnes d'analyse
    analysis_fill = PatternFill(
        start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
    )

    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=column_num)
        # Préserver le style existant (gras, bordures) mais changer la couleur de fond
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            # Si la cellule a déjà une couleur (header bleu, total orange), on la garde
            # Sauf si c'est blanc/transparent
            existing_color = cell.fill.start_color.rgb
            if existing_color in ["00000000", "FFFFFFFF", None]:
                cell.fill = analysis_fill
        else:
            cell.fill = analysis_fill

    # Appliquer la mise en forme conditionnelle pour les variations (colonnes Écart et Variation)
    # Valeurs positives en vert foncé, négatives en rouge
    col_letter = get_column_letter(column_num)
    range_str = f"{col_letter}{start_row}:{col_letter}{end_row}"

    # Règle pour les valeurs positives (vert foncé)
    green_font = Font(color="006100", bold=True)
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator='greaterThan', formula=['0'], font=green_font)
    )

    # Règle pour les valeurs négatives (rouge)
    red_font = Font(color="9C0006", bold=True)
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator='lessThan', formula=['0'], font=red_font)
    )


def adjust_column_widths_suivi(
    ws,
    nb_mois: int,
    col_premier_mois: int,
    col_total_annee: int,
    col_ecart_budget: int,
    col_variation: int,
    col_variation_pct: int,
):
    """
    Ajuste dynamiquement la largeur des colonnes et hauteur des lignes
    pour le suivi d'activité - Mise en forme professionnelle

    Args:
        ws: Worksheet
        nb_mois: Nombre de mois de données
        col_premier_mois: Index de la première colonne de mois
        col_total_annee: Index de la colonne TOTAL ANNEE
        col_ecart_budget: Index de la colonne Écart au budget
        col_variation: Index de la colonne Variation
        col_variation_pct: Index de la colonne Variation %
    """
    # Largeurs optimisées pour éliminer les "###"
    ws.column_dimensions["A"].width = 3  # Colonne vide
    ws.column_dimensions["B"].width = 50  # Libellés (élargi)
    ws.column_dimensions["C"].width = 15  # Rappel N-1 (élargi)
    ws.column_dimensions["D"].width = 15  # Budget (élargi)

    # Largeur pour les colonnes de mois (dynamique) - élargie
    for mois_idx in range(col_premier_mois, col_premier_mois + nb_mois):
        col_letter = get_column_letter(mois_idx)
        ws.column_dimensions[col_letter].width = 14

    # Largeur pour les 4 colonnes d'analyse - bien élargies
    for col_idx in [
        col_total_annee,
        col_ecart_budget,
        col_variation,
        col_variation_pct,
    ]:
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 16

    # Hauteur des lignes minimum 20 pixels pour meilleure lisibilité
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20


def adjust_column_widths(ws):
    """Ajuste automatiquement la largeur des colonnes en tenant compte des formules et des nombres formatés"""
    # Largeurs par défaut optimisées pour le suivi d'activité
    default_widths = {
        "A": 3,  # Colonne vide
        "B": 40,  # Libellés (plus large pour les noms de comptes)
        "C": 12,  # Rappel N-1
        "D": 12,  # Budget
        "E": 12,  # Janvier
        "F": 12,  # Février
        "G": 12,  # Mars
        "H": 12,  # Avril
        "I": 12,  # Mai
        "J": 12,  # Juin
        "K": 12,  # Juillet
        "L": 12,  # Août
        "M": 12,  # Septembre
        "N": 14,  # TOTAL ANNEE (jaune)
        "O": 14,  # Écart au budget (jaune)
        "P": 14,  # Variation (jaune)
        "Q": 14,  # Variation % (jaune)
    }

    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)

        # Utiliser la largeur par défaut si elle existe
        if column_letter in default_widths:
            ws.column_dimensions[column_letter].width = default_widths[column_letter]
        else:
            # Sinon, calculer automatiquement
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        if isinstance(cell.value, str):
                            cell_length = len(str(cell.value))
                        elif isinstance(cell.value, (int, float)):
                            cell_length = len(f"{cell.value:,.0f}") + 2
                        else:
                            cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass

            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def save_workbook(wb: Workbook, output_path: str):
    """Sauvegarde le classeur Excel"""
    logger.info(f"Sauvegarde du classeur: {output_path}")

    try:
        wb.save(output_path)
        logger.info("Classeur sauvegardé avec succès")

    except Exception as e:
        logger.error(f"Erreur lors de la sauvegarde: {e}")
        raise ExcelGenerationError(f"Impossible de sauvegarder le fichier Excel: {e}")
