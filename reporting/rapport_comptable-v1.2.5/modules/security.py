#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module de Sécurité - Protection contre l'utilisation non autorisée

Fonctionnalités:
1. Journalisation complète des activités (audit log)
2. Limitation à 4 rapports par jour par utilisateur
3. Watermarking des rapports Excel générés
"""

import json
import logging
import socket
import getpass
from datetime import datetime
from pathlib import Path
from typing import Tuple
from openpyxl.styles import Font, Alignment


# ============================================================================
# CONFIGURATION
# ============================================================================

DAILY_REPORT_LIMIT = 4  # Maximum 4 rapports par jour


# ============================================================================
# JOURNALISATION (AUDIT LOG)
# ============================================================================

def setup_audit_log() -> Path:
    """
    Configure le système de journalisation des activités

    Returns:
        Path vers le fichier de log
    """
    # Créer le répertoire des logs d'audit
    audit_dir = Path(__file__).parent.parent / "audit_logs"
    audit_dir.mkdir(exist_ok=True)

    # Fichier de log mensuel
    current_month = datetime.now().strftime("%Y-%m")
    log_file = audit_dir / f"audit_{current_month}.log"

    # Configurer le logger spécifique pour l'audit
    audit_logger = logging.getLogger("audit")
    audit_logger.setLevel(logging.INFO)

    # Handler pour le fichier
    handler = logging.FileHandler(log_file, encoding='utf-8')
    handler.setLevel(logging.INFO)

    # Format détaillé
    formatter = logging.Formatter(
        '%(asctime)s | %(levelname)s | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    handler.setFormatter(formatter)

    # Ajouter le handler (éviter les doublons)
    if not audit_logger.handlers:
        audit_logger.addHandler(handler)

    return log_file


def log_session_start(username: str):
    """Enregistre le début d'une session"""
    audit_logger = logging.getLogger("audit")

    hostname = socket.gethostname()

    audit_logger.info(
        f"SESSION_START | User: {username} | Machine: {hostname} | IP: {get_local_ip()}"
    )


def log_report_generation(username: str, client_code: str, gl_file: str,
                          output_excel: str):
    """Enregistre la génération d'un rapport"""
    audit_logger = logging.getLogger("audit")

    audit_logger.info(
        f"REPORT_GENERATED | User: {username} | Client: {client_code} | "
        f"GL: {Path(gl_file).name} | Excel: {Path(output_excel).name}"
    )


def log_limit_exceeded(username: str, current_count: int, limit: int):
    """Enregistre un dépassement de limite"""
    audit_logger = logging.getLogger("audit")

    audit_logger.warning(
        f"LIMIT_EXCEEDED | User: {username} | Count: {current_count} | Limit: {limit}"
    )


# ============================================================================
# LIMITATION D'UTILISATION (4 RAPPORTS/JOUR)
# ============================================================================

def check_daily_limit(username: str, daily_limit: int = DAILY_REPORT_LIMIT) -> Tuple[bool, int, int]:
    """
    Vérifie que l'utilisateur n'a pas dépassé sa limite quotidienne

    Args:
        username: Nom d'utilisateur
        daily_limit: Limite quotidienne de rapports (défaut: 4)

    Returns:
        Tuple (allowed: bool, current_count: int, limit: int)
    """
    logger = logging.getLogger(__name__)

    # Fichier de suivi de l'utilisation
    usage_dir = Path(__file__).parent.parent / "usage_data"
    usage_dir.mkdir(exist_ok=True)
    usage_file = usage_dir / "usage_tracking.json"

    # Charger les données d'utilisation
    if usage_file.exists():
        with open(usage_file, 'r', encoding='utf-8') as f:
            usage_data = json.load(f)
    else:
        usage_data = {}

    # Clé pour le jour en cours
    current_day = datetime.now().strftime("%Y-%m-%d")
    key = f"{username}_{current_day}"

    # Récupérer le compteur
    current_count = usage_data.get(key, 0)

    # Vérifier la limite
    if current_count >= daily_limit:
        logger.error(f"❌ Limite quotidienne dépassée pour {username}: {current_count}/{daily_limit}")
        log_limit_exceeded(username, current_count, daily_limit)
        return False, current_count, daily_limit

    # Incrémenter le compteur
    usage_data[key] = current_count + 1

    # Sauvegarder
    with open(usage_file, 'w', encoding='utf-8') as f:
        json.dump(usage_data, f, indent=2)

    logger.info(f"✅ Utilisation: {username} - {current_count + 1}/{daily_limit} rapports aujourd'hui")

    return True, current_count + 1, daily_limit


def display_usage_stats(username: str):
    """
    Affiche les statistiques d'utilisation pour un utilisateur

    Args:
        username: Nom d'utilisateur
    """
    usage_dir = Path(__file__).parent.parent / "usage_data"
    usage_file = usage_dir / "usage_tracking.json"

    if not usage_file.exists():
        print(f"📊 Aucune utilisation enregistrée pour aujourd'hui")
        return

    with open(usage_file, 'r', encoding='utf-8') as f:
        usage_data = json.load(f)

    # Filtrer les données de cet utilisateur pour aujourd'hui
    current_day = datetime.now().strftime("%Y-%m-%d")
    key = f"{username}_{current_day}"

    count = usage_data.get(key, 0)

    print()
    print("📊 Statistiques d'utilisation:")
    print(f"   • Utilisateur: {username}")
    print(f"   • Date: {datetime.now().strftime('%d/%m/%Y')}")
    print(f"   • Rapports générés aujourd'hui: {count}/{DAILY_REPORT_LIMIT}")
    print(f"   • Rapports restants: {DAILY_REPORT_LIMIT - count}")
    print()


# ============================================================================
# WATERMARKING
# ============================================================================

def add_watermark_to_sheet(ws, company_name: str = "2BN CONSULTING", row_offset: int = 2):
    """
    Ajoute un filigrane (watermark) à une feuille Excel

    Args:
        ws: Worksheet openpyxl
        company_name: Nom de l'entreprise (cabinet)
        row_offset: Décalage à partir de la dernière ligne
    """
    # Ajouter en bas de la feuille
    last_row = ws.max_row + row_offset

    # Informations de génération
    timestamp = datetime.now().strftime("%d/%m/%Y à %H:%M:%S")

    watermark_text = f"📄 Généré par {company_name} le {timestamp}"

    # Ajouter la cellule
    cell = ws.cell(row=last_row, column=1)
    cell.value = watermark_text
    cell.font = Font(size=9, italic=True, color="808080", name="Calibri")
    cell.alignment = Alignment(horizontal="left")

    # Log dans l'audit
    audit_logger = logging.getLogger("audit")
    audit_logger.info(f"WATERMARK_ADDED | Sheet: {ws.title} | Company: {company_name}")


def add_watermark_to_workbook(wb, company_name: str = "2BN CONSULTING"):
    """
    Ajoute un watermark à toutes les feuilles d'un classeur Excel

    Args:
        wb: Workbook openpyxl
        company_name: Nom de l'entreprise (cabinet)
    """
    logger = logging.getLogger(__name__)

    # Parcourir toutes les feuilles
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        add_watermark_to_sheet(ws, company_name, row_offset=2)

    logger.info(f"✅ Watermark '{company_name}' ajouté à {len(wb.sheetnames)} feuille(s)")


# ============================================================================
# UTILITAIRES
# ============================================================================

def get_local_ip() -> str:
    """Récupère l'adresse IP locale"""
    try:
        # Créer un socket pour obtenir l'IP
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "Unknown"


def get_current_username() -> str:
    """
    Récupère le nom d'utilisateur système

    Returns:
        Nom d'utilisateur
    """
    return getpass.getuser()


# ============================================================================
# FONCTION PRINCIPALE DE SÉCURITÉ
# ============================================================================

def security_check() -> Tuple[bool, str]:
    """
    Effectue toutes les vérifications de sécurité

    Returns:
        Tuple (authorized: bool, username: str)
    """
    logger = logging.getLogger(__name__)

    # Récupérer le nom d'utilisateur système
    username = get_current_username()

    print()
    print("=" * 80)
    print("           🔒 VÉRIFICATION DE SÉCURITÉ")
    print("=" * 80)
    print()

    # 1. Configurer l'audit log
    audit_log_file = setup_audit_log()
    logger.info(f"Audit log: {audit_log_file}")

    # 2. Enregistrer le début de session
    log_session_start(username)

    # 3. Vérifier la limite quotidienne
    allowed, current_count, limit = check_daily_limit(username)

    if not allowed:
        print()
        print("=" * 80)
        print("           ❌ LIMITE QUOTIDIENNE DÉPASSÉE")
        print("=" * 80)
        print()
        print(f"Vous avez atteint votre limite de {limit} rapports par jour.")
        print(f"Rapports générés aujourd'hui: {current_count}")
        print()
        print("Veuillez réessayer demain ou contactez l'administrateur.")
        print()
        return False, username

    # Afficher les statistiques
    display_usage_stats(username)

    logger.info("✅ Vérifications de sécurité passées avec succès")
    return True, username
