#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Code de remplacement pour add_suivi_activite_sheet
Ce fichier contient la nouvelle version simplifiée qui utilise les groupes intelligents
"""

def add_suivi_activite_sheet_new(wb, df, client_code=None):
    """
    Version simplifiée qui utilise les groupes intelligents au lieu des mappings statiques
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from modules.data_processor import prepare_suivi_activite_detaille, load_client_mapping

    logger.info("Ajout de la feuille Suivi d'Activité - Approche directe depuis GL")

    # Préparer les données avec regroupement intelligent
    suivi_data = prepare_suivi_activite_detaille(df, client_code=client_code)

    ws = wb.create_sheet("SUIVI ACTIVITE")

    # Extraire l'année et le nom du client
    df_with_date = df[df["date"].notna()]
    if len(df_with_date) > 0:
        annee = df_with_date["date"].dt.year.mode()[0]
    else:
        annee = 2025

    # Déterminer le nom du client pour le titre
    if client_code:
        client_mapping = load_client_mapping(client_code)
        if client_mapping and 'client_name' in client_mapping:
            client_name = client_mapping['client_name']
        else:
            client_name = client_code.upper().replace('_', ' ')
    else:
        client_name = "SUIVI"

    # Détecter les mois présents dans les données
    from modules.excel_generator import detect_available_months
    mois_disponibles = detect_available_months(df)
    nb_mois = len(mois_disponibles)

    # Calculer le nombre total de colonnes
    nb_colonnes_total = 4 + nb_mois + 4

    # Position des colonnes
    col_premier_mois = 5
    col_dernier_mois = col_premier_mois + nb_mois - 1
    col_total_annee = col_dernier_mois + 1
    col_ecart_budget = col_total_annee + 1
    col_variation = col_ecart_budget + 1
    col_variation_pct = col_variation + 1

    logger.info(f"Configuration dynamique: {nb_mois} mois détectés, {nb_colonnes_total} colonnes totales")

    # Ligne 1: Titre principal
    ws.append(["", f"Tableau de suivi {client_name} {annee}"])
    end_col_letter = get_column_letter(nb_colonnes_total)
    ws.merge_cells(f"A1:{end_col_letter}1")

    title_cell = ws["A1"]
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.font = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Lignes vides
    for _ in range(3):
        ws.append([""] * nb_colonnes_total)

    # Ligne 5: En-têtes
    mois_headers = ["", "", "Rappel année N-1", "BUDGET PREVI"]
    mois_headers.extend(mois_disponibles)
    mois_headers.extend(["TOTAL ANNEE", "Ecart au budget", "Variation", "Variation %"])
    ws.append(mois_headers)

    from modules.excel_generator import apply_header_style
    apply_header_style(ws, 5, nb_colonnes_total)

    # Fonction helper pour ajouter une ligne
    def add_data_row(libelle, mois_data, indent_level=0, is_subtotal=False):
        indent = "  " * indent_level
        row_data = ["", f"{indent}{libelle}", 0, 0]

        for mois in mois_disponibles:
            montant = mois_data.get(mois, 0)
            row_data.append(montant if montant != 0 else 0)

        ws.append(row_data)
        current_row = ws.max_row

        # Formules
        col_premier_mois_letter = get_column_letter(col_premier_mois)
        col_dernier_mois_letter = get_column_letter(col_dernier_mois)
        col_total_letter = get_column_letter(col_total_annee)

        total_cell = ws.cell(row=current_row, column=col_total_annee)
        total_cell.value = f"=SUM({col_premier_mois_letter}{current_row}:{col_dernier_mois_letter}{current_row})"
        total_cell.number_format = "# ##0"

        ecart_cell = ws.cell(row=current_row, column=col_ecart_budget)
        ecart_cell.value = f"={col_total_letter}{current_row}-D{current_row}"
        ecart_cell.number_format = "# ##0"

        variation_cell = ws.cell(row=current_row, column=col_variation)
        variation_cell.value = f"={col_total_letter}{current_row}-C{current_row}"
        variation_cell.number_format = "# ##0"

        variation_pct_cell = ws.cell(row=current_row, column=col_variation_pct)
        variation_pct_cell.value = f"=IF(C{current_row}=0,0,({col_total_letter}{current_row}-C{current_row})/C{current_row}*100)"
        variation_pct_cell.number_format = "0.00%"

        # Style
        if is_subtotal:
            for cell in ws[current_row]:
                if cell.column <= nb_colonnes_total:
                    cell.font = Font(bold=True)

        return current_row

    # SECTION 1: CHARGES DE PERSONNEL
    personnel_groups = suivi_data.get('personnel', {}).get('groups', [])
    personnel_rows = []

    if personnel_groups:
        for group in personnel_groups:
            if group.get('is_subtotal'):
                # C'est un groupe avec sous-total
                # D'abord ajouter les entrées individuelles
                for entry in group.get('entries', []):
                    add_data_row(entry['libelle'], entry['mois_data'], indent_level=1)

                # Puis le sous-total
                row_num = add_data_row(group['libelle'], group['mois_data'], indent_level=1, is_subtotal=True)
                personnel_rows.append(row_num)
            else:
                # Entrée simple
                add_data_row(group['libelle'], group['mois_data'], indent_level=1)
                personnel_rows.append(ws.max_row)

    # Total charges de personnel
    total_personnel_row_num = ws.max_row + 1
    total_personnel_row = ["", "Total charges de personnel", 0, 0]

    for col_idx in range(col_premier_mois, col_variation_pct + 1):
        col_letter = get_column_letter(col_idx)
        if personnel_rows:
            formula_parts = [f"{col_letter}{row}" for row in personnel_rows]
            total_personnel_row.append(f"={'+'.join(formula_parts)}")
        else:
            total_personnel_row.append(0)

    ws.append(total_personnel_row)
    from modules.excel_generator import apply_total_style
    apply_total_style(ws, ws.max_row, nb_colonnes_total)

    # SECTION 2: AUTRES CHARGES
    charges_groups = suivi_data.get('charges', {}).get('groups', [])
    charges_rows = []

    if charges_groups:
        for group in charges_groups:
            if group.get('is_subtotal'):
                for entry in group.get('entries', []):
                    add_data_row(entry['libelle'], entry['mois_data'], indent_level=1)
                row_num = add_data_row(group['libelle'], group['mois_data'], indent_level=1, is_subtotal=True)
                charges_rows.append(row_num)
            else:
                add_data_row(group['libelle'], group['mois_data'], indent_level=1)
                charges_rows.append(ws.max_row)

    # Total autres charges
    total_autres_charges_row_num = ws.max_row + 1
    total_autres_row = ["", "Autres charges", 0, 0]

    for col_idx in range(col_premier_mois, col_variation_pct + 1):
        col_letter = get_column_letter(col_idx)
        if charges_rows:
            formula_parts = [f"{col_letter}{row}" for row in charges_rows]
            total_autres_row.append(f"={'+'.join(formula_parts)}")
        else:
            total_autres_row.append(0)

    ws.append(total_autres_row)
    apply_total_style(ws, ws.max_row, nb_colonnes_total)

    # Lignes vides
    for _ in range(3):
        ws.append([""] * nb_colonnes_total)

    # Total charges
    total_charges_row_num = ws.max_row + 1
    total_charges_row = ["", "Total charges",
                        f"=C{total_personnel_row_num}+C{total_autres_charges_row_num}",
                        f"=D{total_personnel_row_num}+D{total_autres_charges_row_num}"]

    for col_idx in range(col_premier_mois, col_variation_pct + 1):
        col_letter = get_column_letter(col_idx)
        total_charges_row.append(f"={col_letter}{total_personnel_row_num}+{col_letter}{total_autres_charges_row_num}")

    ws.append(total_charges_row)
    apply_total_style(ws, ws.max_row, nb_colonnes_total)

    for col in range(1, nb_colonnes_total + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(top=Side(style="medium", color="000000"), bottom=Side(style="thin", color="CCCCCC"))

    # Lignes vides
    for _ in range(2):
        ws.append([""] * nb_colonnes_total)

    # SECTION 3: PRODUITS
    produits_groups = suivi_data.get('produits', {}).get('groups', [])
    produits_rows = []

    if produits_groups:
        for group in produits_groups:
            if group.get('is_subtotal'):
                for entry in group.get('entries', []):
                    add_data_row(entry['libelle'], entry['mois_data'], indent_level=1)
                row_num = add_data_row(group['libelle'], group['mois_data'], indent_level=1, is_subtotal=True)
                produits_rows.append(row_num)
            else:
                add_data_row(group['libelle'], group['mois_data'], indent_level=1)
                produits_rows.append(ws.max_row)

    # Ligne vide
    ws.append([""] * nb_colonnes_total)

    # Total CA
    total_ca_row_num = ws.max_row + 1
    total_ca_row = ["", "CA", 0, 0]

    for col_idx in range(col_premier_mois, col_variation_pct + 1):
        col_letter = get_column_letter(col_idx)
        if produits_rows:
            formula_parts = [f"{col_letter}{row}" for row in produits_rows]
            total_ca_row.append(f"={'+'.join(formula_parts)}")
        else:
            total_ca_row.append(0)

    ws.append(total_ca_row)
    apply_total_style(ws, ws.max_row, nb_colonnes_total)

    # Ligne vide
    ws.append([""] * nb_colonnes_total)

    # RÉSULTAT AVANT IMPÔTS
    resultat_row = ["", "Résultat avant impôts",
                    f"=C{total_ca_row_num}+C{total_charges_row_num}",
                    f"=D{total_ca_row_num}+D{total_charges_row_num}"]

    for col_idx in range(col_premier_mois, col_variation_pct + 1):
        col_letter = get_column_letter(col_idx)
        resultat_row.append(f"={col_letter}{total_ca_row_num}+{col_letter}{total_charges_row_num}")

    ws.append(resultat_row)
    from modules.excel_generator import apply_result_style
    apply_result_style(ws, ws.max_row, nb_colonnes_total)

    # Lignes vides finales
    for _ in range(3):
        ws.append([""] * nb_colonnes_total)

    # Formatage des nombres
    from modules.excel_generator import apply_number_format, apply_yellow_column, adjust_column_widths_suivi
    for row in range(6, ws.max_row + 1):
        for col_idx in range(3, nb_colonnes_total + 1):
            apply_number_format(ws, get_column_letter(col_idx), row, row)

    # Couleur jaune aux colonnes d'analyse
    for col_num in [col_total_annee, col_ecart_budget, col_variation, col_variation_pct]:
        apply_yellow_column(ws, col_num, 5, ws.max_row)

    # Ajuster largeurs
    adjust_column_widths_suivi(ws, nb_mois, col_premier_mois, col_total_annee,
                               col_ecart_budget, col_variation, col_variation_pct)

    # Figer les volets
    ws.freeze_panes = "B6"
